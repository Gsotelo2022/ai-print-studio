from flask import Flask, request, jsonify
from flask_cors import CORS
import psycopg2
import requests
import re
import json
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import os

app = Flask(__name__)
CORS(app)

# ==============================
# CONFIG
# ==============================

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL = "qwen2.5:1.5b"
FAQ_FILE = os.path.join(os.path.dirname(__file__), "faq_sql_ollama.xlsx")
SIMILARITY_THRESHOLD = 0.3  # Umbral mínimo de similitud

# ==============================
# CACHE FAQ
# ==============================

faq_data = None
vectorizer = None
tfidf_matrix = None

def cargar_faq():
    """Carga FAQ del Excel y prepara vectorizador TF-IDF"""
    global faq_data, vectorizer, tfidf_matrix
    
    try:
        if not os.path.exists(FAQ_FILE):
            print(f"❌ Archivo FAQ no encontrado: {FAQ_FILE}")
            return False
        
        df = pd.read_excel(FAQ_FILE)
        faq_data = df.copy()
        
        # Preparar vectorizador TF-IDF
        # Nota: scikit-learn solo soporta 'english' por defecto
        # Usamos None para cargar todas las palabras
        vectorizer = TfidfVectorizer(lowercase=True, stop_words=None)
        tfidf_matrix = vectorizer.fit_transform(faq_data['Pregunta Frecuente'].astype(str))
        
        # Verificar que existe columna 'Tipo'
        if 'Tipo' not in faq_data.columns:
            print(f"⚠️ Columna 'Tipo' no encontrada en FAQ")
        
        print(f"✅ FAQ cargado: {len(faq_data)} preguntas")
        return True
    except Exception as e:
        print(f"❌ Error cargando FAQ: {e}")
        return False

def buscar_pregunta_similar(consulta_usuario):
    """Busca la pregunta más similar en el FAQ usando TF-IDF"""
    global faq_data, vectorizer, tfidf_matrix
    
    if faq_data is None or vectorizer is None:
        return None, None, 0
    
    try:
        # Vectorizar consulta del usuario
        user_vector = vectorizer.transform([consulta_usuario])
        
        # Calcular similitud con todas las preguntas
        similarities = cosine_similarity(user_vector, tfidf_matrix)[0]
        
        # Obtener índice de mayor similitud
        best_idx = np.argmax(similarities)
        best_score = similarities[best_idx]
        
        if best_score < SIMILARITY_THRESHOLD:
            print(f"⚠️ Similitud muy baja ({best_score:.2f}) - debajo del umbral")
            return None, None, best_score
        
        pregunta = faq_data.iloc[best_idx]['Pregunta Frecuente']
        query_sql = faq_data.iloc[best_idx]['Query SQL']
        
        print(f"🎯 Pregunta similar encontrada (similitud: {best_score:.2f})")
        print(f"   Pregunta: {pregunta[:60]}...")
        
        return pregunta, query_sql, best_score
    
    except Exception as e:
        print(f"❌ Error buscando pregunta similar: {e}")
        return None, None, 0

# ==============================
# DB CONNECTION — PostgreSQL
# ==============================

DB_CONFIG = {
    'host': os.getenv('PG_HOST', '127.0.0.1'),
    'port': os.getenv('PG_PORT', '5432'),
    'database': os.getenv('PG_DB', 'PrendeteRock'),
    'user': os.getenv('PG_USER', 'postgres'),
    'password': os.getenv('PG_PASSWORD', 'Pasteldepapas123#')
}

def get_db_connection():
    try:
        return psycopg2.connect(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            database=DB_CONFIG['database'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password']
        )
    except Exception as e:
        print(f"❌ Error conectando a BD PostgreSQL: {e}")
        return None

def obtener_rol_usuario(user_id):
    """Obtiene el tipo de usuario (admin o cliente) desde la BD PostgreSQL"""
    try:
        conn = get_db_connection()
        if not conn:
            print(f"⚠️ No se pudo verificar tipo del usuario {user_id}")
            return "cliente"  # Default a cliente si hay error
        
        cursor = conn.cursor()
        
        # Buscar el campo 'tipo' en la tabla usuarios (PostgreSQL)
        query = "SELECT tipo FROM usuarios WHERE id_usuario = %s"
        
        try:
            cursor.execute(query, (user_id,))
            result = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if result:
                tipo = result[0].lower() if result[0] else "cliente"
                print(f"✅ Tipo de usuario encontrado: {tipo}")
                return tipo
            else:
                print(f"⚠️ Usuario {user_id} no encontrado")
                return "cliente"
        except Exception as e:
            print(f"❌ Error en query: {e}")
            conn.close()
            return "cliente"
        
    except Exception as e:
        print(f"⚠️ Error obteniendo tipo de usuario: {e}")
        return "cliente"  # Default a cliente

# ==============================
# FUNCIONES DB
# ==============================

def ejecutar_query_faq(query_sql, user_id):
    """Ejecuta query del FAQ reemplazando :id_usuario con el user_id
    
    Maneja dos casos:
    - Queries CON :id_usuario → Pasa user_id como parámetro
    - Queries SIN :id_usuario → Ejecuta sin parámetro (queries genéricas/catálogo)
    
    También CORRIGE sintaxis SQL Server (LIMIT → TOP)
    """
    try:
        conn = get_db_connection()
        if conn is None:
            return None, "Error de conexión a la BD"
        
        cursor = conn.cursor()
        
        # Contar placeholders para todos los posibles parámetros
        param_count = query_sql.count(':id_usuario') + query_sql.count(':user_id') + query_sql.count(':id_pedido') + query_sql.count(':id_item') + query_sql.count(':id_variante') + query_sql.count(':numero_orden') + query_sql.count(':codigo') + query_sql.count(':id_pago') + query_sql.count(':id_producto') + query_sql.count(':texto') + query_sql.count(':id_archivo')
        
        # PARA POSTGRESQL: Reemplazar placeholders :param con %s
        query_procesada = query_sql
        query_procesada = query_procesada.replace(':id_usuario', '%s')
        query_procesada = query_procesada.replace(':user_id', '%s')
        query_procesada = query_procesada.replace(':id_pedido', '%s')
        query_procesada = query_procesada.replace(':id_item', '%s')
        query_procesada = query_procesada.replace(':id_variante', '%s')
        query_procesada = query_procesada.replace(':numero_orden', '%s')
        query_procesada = query_procesada.replace(':codigo', '%s')
        query_procesada = query_procesada.replace(':id_pago', '%s')
        query_procesada = query_procesada.replace(':id_producto', '%s')
        query_procesada = query_procesada.replace(':texto', '%s')
        query_procesada = query_procesada.replace(':id_archivo', '%s')
        
        print(f"📊 Ejecutando query: {query_procesada[:100]}...")
        print(f"   Parámetros: {param_count} (user_id: {user_id})")
        
        # Ejecutar con o sin parámetros según corresponda
        if param_count > 0:
            # Query con filtro de usuario - pasar user_id como tupla
            cursor.execute(query_procesada, (user_id,))
        else:
            # Query genérica (sin parámetros)
            print(f"   ⚠️ Query genérica detectada (sin parámetros)")
            cursor.execute(query_procesada)
        
        rows = cursor.fetchall()
        conn.close()
        
        return rows, None
    
    except Exception as e:
        print(f"❌ Error ejecutando query: {e}")
        return None, str(e)

def formatear_resultado(rows, columnas_cursor):
    """Formatea resultados de BD en texto legible"""
    if not rows:
        return "Sin resultados"
    
    try:
        # Obtener nombres de columnas
        col_names = [desc[0] for desc in columnas_cursor.description] if hasattr(columnas_cursor, 'description') else []
        
        if len(rows) == 1 and len(rows[0]) == 1:
            # Un solo valor (ej: COUNT)
            return str(rows[0][0])
        
        # Múltiples filas/columnas
        resultado = []
        for row in rows[:10]:  # Máximo 10 filas
            resultado.append(" | ".join(str(v) for v in row))
        
        return "\n".join(resultado)
    
    except Exception as e:
        print(f"❌ Error formateando: {e}")
        return str(rows)

# ==============================
# (IA DEPRECADA - Ahora usamos similitud TF-IDF)
# ==============================

# def construir_prompt(texto):
#     return f"""... (código antiguo)"""

# def llamar_ollama(prompt):
#     ... (código antiguo)

# def parsear_json(texto):
#     ... (código antiguo)

def buscar_usuario_por_nombre(nombre_busqueda):
    """Busca un usuario por nombre completo o parcial"""
    try:
        conn = get_db_connection()
        if not conn:
            return None
        
        cursor = conn.cursor()
        
        # Buscar usuario cuyo nombre contenga el texto buscado (case-insensitive)
        query = "SELECT id_usuario, nombre FROM usuarios WHERE LOWER(nombre) LIKE LOWER(%s) LIMIT 1"
        cursor.execute(query, (f"%{nombre_busqueda}%",))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if result:
            user_id, nombre = result
            print(f"👤 Usuario encontrado: {nombre} (ID: {user_id})")
            return user_id
        
        print(f"⚠️ Usuario '{nombre_busqueda}' no encontrado")
        return None
        
    except Exception as e:
        print(f"❌ Error buscando usuario: {e}")
        return None

def extraer_nombre_de_pregunta(pregunta):
    """Extrae posibles nombres propios de la pregunta (palabras con capital)"""
    import re
    # Palabras que típicamente no son nombres (preposiciones, artículos, incluso al inicio de frase)
    palabras_ignoradas = {
        'de', 'del', 'la', 'el', 'un', 'una', 'y', 'o', 'que', 
        'tiene', 'tengo', 'cuales', 'cual', 'cuantos', 'cuanta',
        'son', 'los', 'las', 'mis', 'mis', 'estos', 'esta'
    }
    
    # Buscar palabras con capital (potenciales nombres propios)
    # Ignorar la primera palabra si es una de las ignoradas (aunque empiece con Mayúscula)
    palabras = re.findall(r'\b[A-ZÁÉÍÓÚ][a-záéíóú]+\b', pregunta)
    
    # Filtrar palabras ignoradas (case-insensitive) y palabras muy cortas
    nombres_candidatos = [p for p in palabras if p.lower() not in palabras_ignoradas and len(p) > 2]
    
    if nombres_candidatos:
        # Si la pregunta empieza con una palabra capitalizada pero común, la quitamos
        # Ejemplo: "¿Cuales son..." -> Cuales es la primera, pero está en ignoradas.
        # El regex [A-Z] ya captura "Cuales". Si "Cuales" está en candidatos, lo filtramos.
        
        # Intentamos reconstruir un nombre (Nombre Apellido)
        # Solo si los candidatos aparecen consecutivamente o son relevantes
        nombre_completo = ' '.join(nombres_candidatos[:2]) if len(nombres_candidatos) >= 2 else nombres_candidatos[0]
        print(f"🔤 Nombre detectado en pregunta: {nombre_completo}")
        return nombre_completo
    
    return None

def generar_query_con_ollama(mensaje, tipo_usuario="cliente"):
    """Usa Ollama para generar SQL dinámicamente cuando FAQ no tiene respuesta"""
    try:
        print(f"🤖 Generando query con Ollama...")
        
        # Prompt para guiar al modelo en generar SQL seguro
        prompt = f"""Eres un experto en SQL PostgreSQL. Basándote en esta pregunta del usuario, genera UNA ÚNICA query SQL válida.

CONTEXTO:
- Base de datos: PostgreSQL
- Tablas principales: usuarios, pedidos, pedidos_items, productos, archivos_diseno, pagos, cupones
- El usuario es: {tipo_usuario}

RESTRICCIONES:
- La query debe usar ':id_usuario' como placeholder para filtros de usuario
- NUNCA incluir DROP, DELETE, UPDATE, INSERT o comandos peligrosos
- Las queries deben ser SELECT solamente
- Usar LIMIT para limitar resultados

PREGUNTA DEL USUARIO: {mensaje}

Responde SOLO con la query SQL, sin explicaciones.

Ejemplos válidos:
- SELECT COUNT(*) FROM pedidos WHERE id_usuario = :id_usuario
- SELECT * FROM pedidos WHERE id_usuario = :id_usuario ORDER BY fecha_pedido DESC LIMIT 10
"""
        
        # Llamar a Ollama
        response = requests.post(
            OLLAMA_URL,
            json={"prompt": prompt, "model": MODEL, "stream": False},
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"⚠️ Error Ollama: {response.status_code}")
            return None
        
        resultado = response.json()
        query_generada = resultado.get("response", "").strip()
        
        # Limpiar respuesta (a veces Ollama agrega explicaciones)
        if "SELECT" in query_generada.upper():
            # Extraer la línea que empieza con SELECT
            lineas = query_generada.split('\n')
            for linea in lineas:
                if linea.strip().upper().startswith('SELECT'):
                    query_generada = linea.strip()
                    break
        
        print(f"✅ Query generada: {query_generada[:100]}...")
        return query_generada if query_generada else None
        
    except Exception as e:
        print(f"❌ Error generando query con Ollama: {e}")
        return None

def guardar_pregunta_nueva_al_faq(pregunta, query_sql, tipo="CLIENTE"):
    """Guarda una nueva pregunta y query al archivo Excel del FAQ"""
    try:
        from openpyxl import load_workbook
        
        archivo = FAQ_FILE
        if not os.path.exists(archivo):
            print(f"⚠️ Archivo FAQ no existe: {archivo}")
            return False
        
        wb = load_workbook(archivo)
        ws = wb.active
        
        # Encontrar última fila
        ultima_fila = ws.max_row + 1
        
        # Agregar nueva fila
        ws[f'A{ultima_fila}'] = pregunta
        ws[f'B{ultima_fila}'] = query_sql
        ws[f'C{ultima_fila}'] = tipo
        
        wb.save(archivo)
        print(f"✅ Nueva pregunta guardada en FAQ (fila {ultima_fila})")
        
        # Recargar FAQ en memoria
        cargar_faq()
        
        return True
        
    except Exception as e:
        print(f"❌ Error guardando en FAQ: {e}")
        return False

# ==============================
# ENDPOINT
# ==============================

@app.route("/chat", methods=["POST"])
def chat():
    try:
        data = request.get_json()
        mensaje = data.get("mensaje", "").strip()
        user_id = data.get("user_id")

        if not mensaje:
            return jsonify({"respuesta": "Mensaje vacío"}), 400

        if not user_id:
            return jsonify({"respuesta": "Falta user_id"}), 400

        print(f"\n📩 Mensaje usuario: {mensaje}")
        print(f"👤 User ID: {user_id}")
        
        # ========================================
        # OBTENER TIPO DE USUARIO
        # ========================================
        tipo_usuario = obtener_rol_usuario(user_id)
        print(f"👨‍💼 Tipo del usuario: {tipo_usuario}")
        
        # ========================================
        # DETECTAR SI PREGUNTA POR OTRO USUARIO
        # ========================================
        user_id_a_consultar = user_id  # Default: el usuario logueado
        nombre_detectado = extraer_nombre_de_pregunta(mensaje)
        
        if nombre_detectado:
            # Si hay un nombre específico y NO es admin -> rechazar
            if tipo_usuario.lower() != "admin":
                return jsonify({
                    "respuesta": "❌ Los clientes solo pueden consultar sus propios datos.",
                    "debug": "Detectamos que buscas información de otro usuario."
                }), 403
            
            # Si es admin, buscar el usuario
            user_encontrado = buscar_usuario_por_nombre(nombre_detectado)
            if user_encontrado:
                user_id_a_consultar = user_encontrado
                print(f"📌 Se consultará por usuario: {user_encontrado}")
            else:
                return jsonify({
                    "respuesta": f"❌ Usuario '{nombre_detectado}' no encontrado en la BD."
                }), 404

        # ========================================
        # BUSCAR PREGUNTA SIMILAR EN FAQ
        # ========================================
        pregunta_similar, query_sql, similitud = buscar_pregunta_similar(mensaje)
        
        query_generada_dinamicamente = False

        if query_sql is None:
            # Si similitud es muy baja, intentar generar con Ollama
            print(f"⚠️ Similitud insuficiente ({similitud:.2f}). Intentando generar query...")
            
            query_sql = generar_query_con_ollama(mensaje, tipo_usuario)
            
            if query_sql is None:
                return jsonify({
                    "respuesta": "❌ No conseguí una respuesta cercana a tu consulta y tampoco pude generar una query. Intenta reformular. 🤔",
                    "debug": {"similitud_max": float(similitud)}
                })
            
            pregunta_similar = mensaje
            query_generada_dinamicamente = True
            print(f"🆕 Query generada dinámicamente por IA")

        # ========================================
        # VALIDAR PERMISOS (CLIENTE vs ADMIN)
        # ========================================
        tipo_pregunta = "CLIENTE"  # Default
        if faq_data is not None and 'Tipo' in faq_data.columns and not query_generada_dinamicamente:
            # Buscar índice de la pregunta similar
            for idx, row in faq_data.iterrows():
                if row['Pregunta Frecuente'].lower() == pregunta_similar.lower():
                    tipo_pregunta = row['Tipo']
                    break
        
        # Si fue generada por IA, determinar tipo según contenido
        if query_generada_dinamicamente:
            # Si pregunta por otro usuario y se llegó aquí, es admin
            if user_id_a_consultar != user_id:
                tipo_pregunta = "ADMIN"
            else:
                tipo_pregunta = "CLIENTE"
        
        print(f"🔐 Tipo de pregunta: {tipo_pregunta}")

        # ========================================
        # EJECUTAR QUERY SQL
        # ========================================
        rows, error = ejecutar_query_faq(query_sql, user_id_a_consultar)

        if error:
            return jsonify({
                "respuesta": f"❌ Error al consultar: {error}",
                "pregunta_interpretada": pregunta_similar
            }), 500

        if not rows:
            return jsonify({
                "respuesta": "ℹ️ No hay datos para mostrar.",
                "pregunta_interpretada": pregunta_similar,
                "tipo_consulta": tipo_pregunta
            })

        # ========================================
        # FORMATEAR RESPUESTA
        # ========================================
        resultado_formateado = formatear_resultado(rows, None)
        
        # ========================================
        # GUARDAR QUERY GENERADA EN FAQ
        # ========================================
        if query_generada_dinamicamente:
            guardar_pregunta_nueva_al_faq(pregunta_similar, query_sql, tipo_pregunta)

        return jsonify({
            "respuesta": resultado_formateado,
            "pregunta_interpretada": pregunta_similar,
            "similitud": float(similitud) if not query_generada_dinamicamente else 1.0,
            "tipo_consulta": tipo_pregunta,
            "generada_dinamicamente": query_generada_dinamicamente
        })

    except Exception as e:
        print(f"❌ Error general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"respuesta": f"Error: {str(e)}"}), 500

# ==============================
# HEALTH
# ==============================

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "agente": "pedidos",
        "modelo": MODEL
    })

# ==============================
# RUN
# ==============================

if __name__ == "__main__":
    print("\n" + "="*60)
    print("🚀 INICIANDO AGENTE DE PEDIDOS")
    print("="*60)
    
    # Cargar FAQ
    if cargar_faq():
        print("✅ Sistema listo para consultas")
    else:
        print("⚠️ Advertencia: FAQ no disponible")
    
    print("🌐 Escuchando en http://localhost:5005")
    print("="*60 + "\n")
    
    app.run(port=5005, debug=True)